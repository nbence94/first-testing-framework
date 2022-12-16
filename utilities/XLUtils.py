import openpyxl


def get_row_count(file, sheet_name):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheet_name)
    return sheet.max_row


def get_column_count(file, sheet_name):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheet_name)
    return sheet.max_column


def read_data(file, sheet_name, row_num, column_no):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheet_name)
    return sheet.cell(row=row_num, column=column_no).value


def write_data(file, sheet_name, row_num, column_no, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheet_name)
    sheet.cell(row=row_num, column=column_no).value = data
    workbook.save(file)
